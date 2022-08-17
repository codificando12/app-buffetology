from asyncio import exceptions
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import re
import time
from openpyxl import Workbook 
from openpyxl.styles import Font
import openpyxl

def convert_unicode(unicode):
    try: 
        if unicode.startswith("\u202a−"):
            unicode = unicode.encode("ascii", "ignore")
            unicode = float(unicode)
            unicode = -unicode
            return unicode
        else:
            unicode = unicode.encode("ascii", "ignore")
            unicode = float(unicode)
            return unicode
    except :
        return "N/A"
    
workbook = openpyxl.load_workbook("acciones.xlsx")
ws_more_15 = workbook["more_15"]
ws_more_10 = workbook["more_10"]
ws_more_5 = workbook["more_5"]
ws_more_0 = workbook["more_0"]
ws_less_0 = workbook["less_0"]
ws_stock_01 = workbook["stock_01"]
PATH = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(PATH)
#Step 1: open stocks tabs
driver.get("https://www.tradingview.com/markets/stocks-china/market-movers-all-stocks/")
wait = WebDriverWait(driver, 10)

#save the main tab in a variable
original_windows = driver.current_window_handle

#check that there is only one tab open
assert len(driver.window_handles) == 1

actions = ActionChains(driver)

#I creat a list to try it
stock_numbers = list()
stock_01 = ""

for click_button in range(210):
    try:
        time.sleep(2)
        load_button = driver.find_element(By.CLASS_NAME, "loadButton-59hnCnPW")
        actions.click(load_button)
        actions.perform()
    except:
        break
    # except:
    #     load_button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable(By.CLASS_NAME, "loadButton-59hnCnPW"))
    #     actions.click(load_button)
    #     actions.perform()
        
        
all_stocks = driver.find_element(By.CLASS_NAME, "js-screener-markets-page-init-ssr")
all_stocks = all_stocks.text.split()
    
for quotes in all_stocks:
    if re.findall("^.*([0-9]{6}).*$", quotes):
        stock_numbers.append(quotes)
        if quotes.startswith("688"):
            break
    
        
            
print(stock_numbers)
#lista that will save the shares numbers depending if the return rate is more than
# more_15_num = list()
# more_10_num = list()
# more_5_num = list()
# more_0_num = list()
# less_0_num = list()

#return rate list
# more_15_ret = list()
# more_10_ret = list()
# more_5_ret = list()
# more_0_ret = list()
# less_0_ret = list()


#this loop will go through the list elements and click them
for click in stock_numbers:
    shares = driver.find_element(By.LINK_TEXT, click)
    actions.click(shares)
    actions.perform()
    wait.until(EC.number_of_windows_to_be(2))
    
    # chane tab loop
    for window_handle in driver.window_handles:
        if window_handle != original_windows:
            driver.switch_to.window(window_handle)
            break
    
    
    historical_eps = list()
    
    # click on the financials reports
    try:
        time.sleep(5)
        financial = driver.find_element(By.LINK_TEXT, "Financials")
        actions.click(financial)
        actions.perform()
        
    except:
        time.sleep(15)
        driver.refresh()
        time.sleep(5)
        financial = driver.find_element(By.LINK_TEXT, "Financials")
        actions.click(financial)
        actions.perform()
        
    #class for the share number: tv-symbol-header__second-line--text
    time.sleep(5)
    stock_price = driver.find_element(By.XPATH, '//div[@class="tv-symbol-price-quote__value js-symbol-last"]/span[1]')
    print("The stock price is: " + stock_price.text)
    #if it can't get the stock price, it will give 0.1 and save in another list to do a manual check.
    try: 
        stock_price = float(stock_price.text)
    except:
        stock_price = 0.1
        if stock_price == 0.1:
            stock_01 = click
    statements = driver.find_element(By.LINK_TEXT, "Statements").click()
    income_statement = driver.find_element(By.LINK_TEXT, "Income statement").click()
    time.sleep(5)
    
    try:
        last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[7]/div[1]/div[1]')
        last_eps = last_eps.text
        print("The last EPS is: " + last_eps)
        
        last_eps = convert_unicode(last_eps)
        
        #append historical EPS
        
        
        first_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
        first_eps = convert_unicode(first_eps.text)
        historical_eps.append(first_eps)
        second_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
        second_eps = convert_unicode(second_eps.text)
        historical_eps.append(second_eps)
        thrid_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[3]/div[1]/div[1]')
        thrid_eps = convert_unicode(thrid_eps.text)
        historical_eps.append(thrid_eps)
        fourth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[4]/div[1]/div[1]')
        fourth_eps = convert_unicode(fourth_eps.text)
        historical_eps.append(fourth_eps)
        fifth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[5]/div[1]/div[1]')
        fifth_eps = convert_unicode(fifth_eps.text)
        historical_eps.append(fifth_eps)
        sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[6]/div[1]/div[1]')
        sixth_eps = convert_unicode(sixth_eps.text)
        historical_eps.append(sixth_eps)
        historical_eps.append(last_eps)
    except:
        try:
            last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[6]/div[1]/div[1]')
            last_eps = last_eps.text
            print("The last EPS is: " + last_eps)
            
            last_eps = convert_unicode(last_eps)
            
            #append historical EPS
            first_eps = 0
            historical_eps.append(first_eps)
            second_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
            second_eps = convert_unicode(second_eps.text)
            historical_eps.append(second_eps)
            thrid_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
            thrid_eps = convert_unicode(thrid_eps.text)
            historical_eps.append(thrid_eps)
            fourth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[3]/div[1]/div[1]')
            fourth_eps = convert_unicode(fourth_eps.text)
            historical_eps.append(fourth_eps)
            fifth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[4]/div[1]/div[1]')
            fifth_eps = convert_unicode(fifth_eps.text)
            historical_eps.append(fifth_eps)
            sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[5]/div[1]/div[1]')
            sixth_eps = convert_unicode(sixth_eps.text)
            historical_eps.append(sixth_eps)
            historical_eps.append(last_eps)
        except:
            
            try:
                last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[5]/div[1]/div[1]')
                last_eps = last_eps.text
                print("The last EPS is: " + last_eps)
                
                last_eps = convert_unicode(last_eps)
                
                #append historical EPS
                first_eps = 0
                historical_eps.append(first_eps)
                second_eps = 0
                historical_eps.append(second_eps)
                thrid_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
                thrid_eps = convert_unicode(thrid_eps.text)
                historical_eps.append(thrid_eps)
                fourth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
                fourth_eps = convert_unicode(fourth_eps.text)
                historical_eps.append(fourth_eps)
                fifth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[3]/div[1]/div[1]')
                fifth_eps = convert_unicode(fifth_eps.text)
                historical_eps.append(fifth_eps)
                sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[4]/div[1]/div[1]')
                sixth_eps = convert_unicode(sixth_eps.text)
                historical_eps.append(sixth_eps)
                historical_eps.append(last_eps)
            except:
                try:
                    last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[4]/div[1]/div[1]')
                    last_eps = last_eps.text
                    print("The last EPS is: " + last_eps)
                    
                    last_eps = convert_unicode(last_eps)
                    
                    #append historical EPS
                    first_eps = 0
                    historical_eps.append(first_eps)
                    second_eps = 0
                    historical_eps.append(second_eps)
                    thrid_eps = 0
                    historical_eps.append(thrid_eps)
                    fourth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
                    fourth_eps = convert_unicode(fourth_eps.text)
                    historical_eps.append(fourth_eps)
                    fifth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
                    fifth_eps = convert_unicode(fifth_eps.text)
                    historical_eps.append(fifth_eps)
                    sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[3]/div[1]/div[1]')
                    sixth_eps = convert_unicode(sixth_eps.text)
                    historical_eps.append(sixth_eps)
                    historical_eps.append(last_eps)
                except:
                    try:
                        last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[3]/div[1]/div[1]')
                        last_eps = last_eps.text
                        print("The last EPS is: " + last_eps)
                        
                        last_eps = convert_unicode(last_eps)
                        
                        #append historical EPS
                        first_eps = 0
                        historical_eps.append(first_eps)
                        second_eps = 0
                        historical_eps.append(second_eps)
                        thrid_eps = 0
                        historical_eps.append(thrid_eps)
                        fourth_eps = 0
                        historical_eps.append(fourth_eps)
                        fifth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
                        fifth_eps = convert_unicode(fifth_eps.text)
                        historical_eps.append(fifth_eps)
                        sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
                        sixth_eps = convert_unicode(sixth_eps.text)
                        historical_eps.append(sixth_eps)
                        historical_eps.append(last_eps)
                    except:
                        try:
                            last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[2]/div[1]/div[1]')
                            last_eps = last_eps.text
                            print("The last EPS is: " + last_eps)
                            
                            last_eps = convert_unicode(last_eps)
                            
                            #append historical EPS
                            first_eps = 0
                            historical_eps.append(first_eps)
                            second_eps = 0
                            historical_eps.append(second_eps)
                            thrid_eps = 0
                            historical_eps.append(thrid_eps)
                            fourth_eps = 0
                            historical_eps.append(fourth_eps)
                            fifth_eps = 0
                            historical_eps.append(fifth_eps)
                            sixth_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
                            sixth_eps = convert_unicode(sixth_eps.text)
                            historical_eps.append(sixth_eps)
                            historical_eps.append(last_eps)
                        except:
                            
                            last_eps = driver.find_element(By.XPATH, '//div[@class="container-Odj0lQp6"]/div[20]/div[5]/div[1]/div[1]/div[1]')
                            last_eps = last_eps.text
                            print("The last EPS is: " + last_eps)
                                
                            last_eps = convert_unicode(last_eps)
                                
                            #append historical EPS
                            first_eps = 0
                            historical_eps.append(first_eps)
                            second_eps = 0
                            historical_eps.append(second_eps)
                            thrid_eps = 0
                            historical_eps.append(thrid_eps)
                            fourth_eps = 0
                            historical_eps.append(fourth_eps)
                            fifth_eps = 0
                            historical_eps.append(fifth_eps)
                            sixth_eps = 0
                            historical_eps.append(sixth_eps)
                            historical_eps.append(last_eps)

    
    # get the EPS and calculate the return rate
    try:
        return_rate = round((last_eps / stock_price) * 100, 2)
    except:
        return_rate = 0
    
    if return_rate >= 15:
        share_name = click
        # more_15_num.append(share_name)
        # more_15_ret.append(return_rate)
        ws_more_15.append({"A": share_name, "B": return_rate, "C": historical_eps[0], "D": historical_eps[1], "E": historical_eps[2], "F": historical_eps[3], "G": historical_eps[4], "H": historical_eps[5], "I": historical_eps[6]})
    elif return_rate >= 10:
        share_name = click
        # more_10_num.append(share_name)
        # more_10_ret.append(return_rate)
        ws_more_10.append({"A": share_name, "B": return_rate,"C": historical_eps[0], "D": historical_eps[1], "E": historical_eps[2], "F": historical_eps[3], "G": historical_eps[4], "H": historical_eps[5], "I": historical_eps[6]})
    elif return_rate >= 5:
        share_name = click
        # more_5_num.append(share_name)
        # more_5_ret.append(return_rate)
        ws_more_5.append({"A": share_name, "B": return_rate,"C": historical_eps[0], "D": historical_eps[1], "E": historical_eps[2], "F": historical_eps[3], "G": historical_eps[4], "H": historical_eps[5], "I": historical_eps[6]})
    elif return_rate >= 0:
        share_name = click
        # more_0_num.append(share_name)
        # more_0_ret.append(return_rate)
        ws_more_0.append({"A": share_name, "B": return_rate,"C": historical_eps[0], "D": historical_eps[1], "E": historical_eps[2], "F": historical_eps[3], "G": historical_eps[4], "H": historical_eps[5], "I": historical_eps[6]})
    elif return_rate <= 0:
        share_name = click
        # less_0_num.append(share_name)
        # less_0_ret.append(return_rate)
        ws_less_0.append({"A": share_name, "B": return_rate,"C": historical_eps[0], "D": historical_eps[1], "E": historical_eps[2], "F": historical_eps[3], "G": historical_eps[4], "H": historical_eps[5], "I": historical_eps[6]})
    print("the return rate for " + click + " is: " + str(return_rate) + "%")

    historical_eps = list()
        
    ws_stock_01.append(stock_01)
    workbook.save("acciones.xlsx")
    
    driver.close()
    
    
    #return to the main tab to keep going through the list.
    driver.switch_to.window(original_windows)


#share number append to excel file 
# share_col = 1
# return_col = 2
# row = 2
# for i, value in enumerate(more_15_num, start=row):
#     ws_more_15.cell(row=i, column=share_col).value = value
# for i, value in enumerate(more_15_ret, start=row):
#     ws_more_15.cell(row=i, column=return_col).value = value
    
# for a, value in enumerate(more_10_num, start=row):
#     ws_more_10.cell(row=a, column=share_col).value = value
# for i, value in enumerate(more_10_ret, start=row):
#     ws_more_10.cell(row=i, column=return_col).value = value

# for a, value in enumerate(more_5_num, start=row):
#     ws_more_5.cell(row=a, column=share_col).value = value
# for i, value in enumerate(more_5_ret, start=row):
#     ws_more_5.cell(row=i, column=return_col).value = value
    
# for a, value in enumerate(more_0_num, start=row):
#     ws_more_0.cell(row=a, column=share_col).value = value
# for i, value in enumerate(more_0_ret, start=row):
#     ws_more_0.cell(row=i, column=return_col).value = value
    
# for a, value in enumerate(less_0_num, start=row):
#     ws_less_0.cell(row=a, column=share_col).value = value
# for i, value in enumerate(less_0_ret, start=row):
#     ws_less_0.cell(row=i, column=return_col).value = value
    
# ws_stock_01.append(stock_01)
        
driver.quit()  



    