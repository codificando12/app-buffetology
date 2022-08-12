from venv import create
from numpy import save
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "more_15"
more_10 = create

lst = [0, 1, 2, 3, 4 , 5 , 6]
lst2 = [7, 8, 9, 10, 11, 12]
ws.append({"C":lst[0], "D":lst[1], "E":lst[2], "F":lst[3]})
ws.append({"C":lst[0], "D":lst[1], "E":lst[2], "F":lst[3]})

# for i, value in enumerate(lst, start=2):
#     ws.cell(row=3, column=i).value = value
# for i, value in enumerate(lst2, start=2):
#     ws.cell(row=3, column=i).value = value
wb.save("prueba.xlsx")